#!/usr/bin/env python3
"""Audit TataDiet V4 data and produce immutable V5 phase-1 base snapshots.

This script is intentionally read-only with respect to the source workbook. It:
- validates the ingredient catalog;
- parses every meal ingredient line into structured data;
- reconciles parsed meal ingredients with the workbook's daily detail sheet;
- recalculates nutrients for structured meals;
- freezes ingredients, recipe families/versions and the 180-day plan template;
- writes machine-readable audit reports used by later V5 phases.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import build_site  # noqa: E402

SOURCE_XLSX = ROOT / "source_data" / "Piano_alimentare_revisionato_6_mesi_fibra_moderata.xlsx"
SOURCE_PDF = ROOT / "source_data" / "Piano_alimentare_revisionato_6_mesi_fibra_moderata.pdf"
AUDIT_DIR = ROOT / "v5_audit"
BASE_DIR = ROOT / "v5_data" / "base"
DATASET_VERSION = "tatadiet-base-v1"
PHASE_VERSION = "5.0.0-alpha.1-phase1"
CORE_NUTRIENTS = ("kcal", "protein", "carbs", "fat", "fiber")

TOKEN_RE = re.compile(
    r"^(.*?)[\s\u00a0]+(\d+(?:[.,]\d+)?)\s*(g|ml|uova?|porzioni?|unit[aà])(?:\s*\(([^)]*)\))?$",
    re.IGNORECASE,
)

STATE_PATTERNS = (
    ("drained", re.compile(r"sgocciolat", re.IGNORECASE)),
    ("cooked", re.compile(r"cott[oaie]", re.IGNORECASE)),
    ("dry", re.compile(r"\bsecc[oaie]\b|\bsecchi\b", re.IGNORECASE)),
    ("raw", re.compile(r"crud[oaie]", re.IGNORECASE)),
    ("prepared", re.compile(r"arrost|arrosto", re.IGNORECASE)),
)

SPECIAL_ALIASES = {
    "uovo": "uovo",
    "spinaci ben cotti": "spinaci",
    "banana matura": "banana",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def short_hash(value: Any, length: int = 16) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()[:length]


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"\([^)]*\)", " ", text.casefold())
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def infer_state(name: str) -> tuple[str, str, bool]:
    for state, pattern in STATE_PATTERNS:
        if pattern.search(name):
            return state, "name_inference", False
    return "unspecified", "not_declared", True


def load_catalog(workbook) -> tuple[dict[str, dict[str, Any]], dict[str, str], list[dict[str, Any]]]:
    ws = workbook["Ingredienti"]
    catalog: dict[str, dict[str, Any]] = {}
    name_to_code: dict[str, str] = {}
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if not row[0]:
            continue
        code = str(row[0]).strip()
        name = str(row[1]).strip()
        record = {
            "code": code,
            "name": name,
            "category": str(row[2]).strip(),
            "unit": str(row[3]).strip().lower(),
            "kcal": float(row[4]),
            "protein": float(row[5]),
            "carbs": float(row[6]),
            "fat": float(row[7]),
            "fiber": float(row[8]),
            "source_row": row_number,
        }
        catalog[code] = record
        name_to_code[normalize_text(name)] = code
        rows.append(record)
    for alias, code in SPECIAL_ALIASES.items():
        name_to_code[normalize_text(alias)] = code
    return catalog, name_to_code, rows


def parse_ingredient_token(
    token: str,
    catalog: dict[str, dict[str, Any]],
    name_to_code: dict[str, str],
) -> dict[str, Any]:
    token = token.strip()
    label: str
    quantity: float
    source_unit: str
    note = ""

    special = re.fullmatch(r"(\d+(?:[.,]\d+)?)\s+uov[oa](?:\s+(sodo))?", token, re.IGNORECASE)
    if special:
        label = "uovo"
        quantity = float(special.group(1).replace(",", "."))
        source_unit = "piece"
        note = special.group(2) or ""
    else:
        match = TOKEN_RE.fullmatch(token)
        if not match:
            raise ValueError(f"Token ingrediente non riconosciuto: {token!r}")
        label = match.group(1).strip()
        quantity = float(match.group(2).replace(",", "."))
        source_unit_raw = match.group(3).casefold()
        note = (match.group(4) or "").strip()
        if source_unit_raw.startswith("uov"):
            source_unit = "piece"
        elif source_unit_raw.startswith("porzion"):
            source_unit = "portion"
        elif source_unit_raw.startswith("unit"):
            source_unit = "piece"
        else:
            source_unit = source_unit_raw

    code = name_to_code.get(normalize_text(label))
    if not code:
        raise KeyError(f"Nessun codice catalogo per etichetta {label!r} (token {token!r})")
    ingredient = catalog[code]
    base_unit = ingredient["unit"]
    base_quantity = quantity
    conversion_id = None
    if code == "uovo" and source_unit == "piece":
        base_quantity = quantity * 50.0
        base_unit = "g"
        conversion_id = "base:conversion:uovo-piece-50g"
    elif source_unit != base_unit:
        raise ValueError(
            f"Unità incompatibile per {code}: token={source_unit!r}, catalogo={base_unit!r}, testo={token!r}"
        )

    return {
        "ingredient_id": f"base:ingredient:{code}",
        "ingredient_revision_id": f"base:ingredient-revision:{code}@1",
        "ingredient_code": code,
        "label": label,
        "quantity": quantity,
        "unit": source_unit,
        "base_quantity": round(base_quantity, 6),
        "base_unit": base_unit,
        "conversion_id": conversion_id,
        "preparation_note": note or None,
        "source_text": token,
    }


def calculate_nutrients(lines: list[dict[str, Any]], catalog: dict[str, dict[str, Any]]) -> dict[str, float]:
    totals = {key: 0.0 for key in CORE_NUTRIENTS}
    for line in lines:
        ingredient = catalog[line["ingredient_code"]]
        factor = float(line["base_quantity"]) / 100.0
        for key in CORE_NUTRIENTS:
            totals[key] += factor * float(ingredient[key])
    return {key: round(value, 6) for key, value in totals.items()}




def public_nutrients(values: dict[str, float]) -> dict[str, float]:
    return {
        "energy_kcal": round(float(values["kcal"]), 6),
        "protein_g": round(float(values["protein"]), 6),
        "carbohydrate_g": round(float(values["carbs"]), 6),
        "fat_g": round(float(values["fat"]), 6),
        "fiber_g": round(float(values["fiber"]), 6),
    }

def source_nutrients(meal: dict[str, Any]) -> dict[str, float]:
    return {key: round(float(meal[key]), 6) for key in CORE_NUTRIENTS}


def round_half_up(value: float, digits: int = 1) -> Decimal:
    quantum = Decimal("1").scaleb(-digits)
    return Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP)


def rounded_match(left: dict[str, float], right: dict[str, float], digits: int = 1) -> bool:
    # The workbook stores one-decimal display values produced from binary floats.
    # A source value is therefore considered consistent when it is within half of
    # the displayed unit (0.05 at one decimal), with a tiny numerical margin.
    tolerance = (0.5 * (10 ** (-digits))) + 1e-6
    return all(abs(float(left[key]) - float(right[key])) <= tolerance for key in CORE_NUTRIENTS)


def stable_source_timestamp(workbook) -> str:
    """Return a deterministic timestamp derived from workbook metadata.

    Base seed files must be byte-for-byte reproducible. Using the audit execution
    time in those files would change every hash on every run, so the source
    workbook's modified/created timestamp is used instead.
    """
    value = workbook.properties.modified or workbook.properties.created
    if value is None:
        return "1970-01-01T00:00:00+00:00"
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).replace(microsecond=0).isoformat()


def audit_workbook_formulas(path: Path) -> dict[str, Any]:
    """Verify that formulas have cached values and no cached Excel errors."""
    formula_book = load_workbook(path, data_only=False, read_only=False)
    value_book = load_workbook(path, data_only=True, read_only=False)
    per_sheet: dict[str, Any] = {}
    formula_count = 0
    missing_cached_values = 0
    cached_errors: list[dict[str, str]] = []
    for sheet_name in formula_book.sheetnames:
        formula_ws = formula_book[sheet_name]
        value_ws = value_book[sheet_name]
        sheet_count = 0
        sheet_missing = 0
        sheet_errors: list[dict[str, str]] = []
        for row in formula_ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                sheet_count += 1
                cached = value_ws[cell.coordinate].value
                if cached is None:
                    sheet_missing += 1
                if isinstance(cached, str) and cached.startswith("#"):
                    sheet_errors.append({"cell": cell.coordinate, "value": cached})
        if sheet_count:
            per_sheet[sheet_name] = {
                "formula_count": sheet_count,
                "missing_cached_values": sheet_missing,
                "cached_errors": sheet_errors,
            }
        formula_count += sheet_count
        missing_cached_values += sheet_missing
        cached_errors.extend({"sheet": sheet_name, **item} for item in sheet_errors)
    return {
        "formula_count": formula_count,
        "missing_cached_values": missing_cached_values,
        "cached_errors": cached_errors,
        "per_sheet": per_sheet,
    }


def scan_static_dependencies() -> dict[str, Any]:
    # Scan each source tree once; nested roots would duplicate references.
    targets = [ROOT / "static", ROOT / "templates", ROOT / "scripts"]
    pattern = re.compile(r"data/([a-z0-9_-]+\.json)", re.IGNORECASE)
    references: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen: set[tuple[str, int, str]] = set()
    for target in targets:
        files = [target] if target.is_file() else list(target.rglob("*"))
        for path in files:
            if not path.is_file() or path.suffix.lower() not in {".js", ".py", ".html"}:
                continue
            try:
                lines = path.read_text(encoding="utf-8").splitlines()
            except UnicodeDecodeError:
                continue
            for line_number, line in enumerate(lines, start=1):
                for match in pattern.finditer(line):
                    relative = path.relative_to(ROOT).as_posix()
                    key = (relative, line_number, match.group(1))
                    if key in seen:
                        continue
                    seen.add(key)
                    references[match.group(1)].append(
                        {
                            "file": relative,
                            "line": line_number,
                            "context": line.strip()[:240],
                        }
                    )
    return {
        "generated_at": now_iso(),
        "runtime_and_build_references": dict(sorted(references.items())),
        "reference_count": sum(len(items) for items in references.values()),
        "v5_requirement": (
            "Ogni modulo operativo deve leggere il piano effettivo (base + dati IndexedDB) attraverso un unico repository, "
            "non accedere direttamente ai JSON statici quando esistono personalizzazioni."
        ),
    }


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BASE_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    monthly_summary, overall_summary = build_site.parse_summary(workbook)
    cycles, days, current_recipes = build_site.parse_plan(workbook, monthly_summary)
    ingredient_days, usage_catalog = build_site.parse_ingredient_details(workbook)
    catalog, name_to_code, catalog_rows = load_catalog(workbook)

    parsed_day_totals: dict[int, dict[tuple[str, str], float]] = defaultdict(lambda: defaultdict(float))
    alias_labels: dict[str, set[str]] = defaultdict(set)
    recipe_families: dict[str, dict[str, Any]] = {}
    recipe_versions: dict[str, dict[str, Any]] = {}
    version_occurrences: dict[str, list[str]] = defaultdict(list)
    plan_days: list[dict[str, Any]] = []
    structured_meals = 0
    estimated_meals = 0
    ingredient_line_count = 0
    structured_nutrient_matches = 0
    nutrient_absolute_differences: dict[str, list[float]] = defaultdict(list)
    parse_errors: list[str] = []

    for day in days:
        plan_day = {
            "id": f"base:plan-day:{day['global_day']:03d}",
            "base_global_day": day["global_day"],
            "cycle": day["cycle"],
            "month": day["month"],
            "variant": day["variant"],
            "day_in_variant": day["day_in_variant"],
            "day_type": day["d_code"],
            "shift": {
                "name": day["shift_name"],
                "description": day["shift_hours"],
                "start_local": "08:00" if day["d_code"] == "D1" else ("20:00" if day["d_code"] == "D2" else None),
                "end_local": "20:00" if day["d_code"] == "D1" else ("08:00" if day["d_code"] == "D2" else None),
                "end_day_offset": 1 if day["d_code"] == "D2" else 0,
            },
            "flexible": bool(day["flexible"]),
            "source_total": public_nutrients({key: round(float(day["total"][key]), 6) for key in CORE_NUTRIENTS}),
            "meals": [],
            "immutable": True,
        }

        for sequence, meal in enumerate(day["meals"], start=1):
            occurrence_id = f"base:meal:{day['global_day']:03d}:{sequence:02d}"
            family_id = f"base:recipe:{meal['recipe_slug']}"
            family = recipe_families.setdefault(
                family_id,
                {
                    "id": family_id,
                    "slug": meal["recipe_slug"],
                    "title": meal["title"],
                    "description": None,
                    "origin": "base",
                    "immutable": True,
                    "status": "active",
                    "meal_types": set(),
                    "cuisines": set(),
                    "version_ids": set(),
                    "instructions_status": "missing",
                },
            )
            family["meal_types"].add(meal["meal_type"])
            family["cuisines"].add(meal["cuisine"])

            tokens = [token.strip() for token in str(meal.get("ingredients") or "").split(";") if token.strip()]
            lines: list[dict[str, Any]] = []
            for token in tokens:
                try:
                    line = parse_ingredient_token(token, catalog, name_to_code)
                except (ValueError, KeyError) as exc:
                    parse_errors.append(f"D{day['global_day']} {meal['time']} {meal['title']}: {exc}")
                    continue
                lines.append(line)
                ingredient_line_count += 1
                parsed_day_totals[day["global_day"]][(line["ingredient_code"], line["base_unit"])] += float(line["base_quantity"])
                alias_labels[line["ingredient_code"]].add(line["label"])

            source_values = source_nutrients(meal)
            if lines:
                structured_meals += 1
                calculated = calculate_nutrients(lines, catalog)
                for nutrient_key in CORE_NUTRIENTS:
                    nutrient_absolute_differences[nutrient_key].append(
                        abs(float(source_values[nutrient_key]) - float(calculated[nutrient_key]))
                    )
                composition_status = "structured"
                nutrition_mode = "calculated_from_ingredients"
                nutrient_match = rounded_match(source_values, calculated)
                if nutrient_match:
                    structured_nutrient_matches += 1
            else:
                estimated_meals += 1
                calculated = None
                composition_status = "unstructured_estimate"
                nutrition_mode = "manual_estimate"
                nutrient_match = None

            version_identity = {
                "ingredients_text": str(meal.get("ingredients") or ""),
                "source_kcal_rounded": round(float(meal["kcal"]), 1),
                "prep_minutes": int(meal["prep_minutes"]),
                "prepare_ahead": meal["prepare_ahead"],
                "cold": meal["cold"],
                "reheat": meal["reheat"],
                "fridge": meal["fridge"],
                "cuisine": meal["cuisine"],
                "spices": meal["spices"],
            }
            version_hash = short_hash(version_identity)
            version_id = f"base:recipe-version:{meal['recipe_slug']}:{version_hash}"
            family["version_ids"].add(version_id)

            if version_id not in recipe_versions:
                recipe_versions[version_id] = {
                    "id": version_id,
                    "recipe_id": family_id,
                    "revision": 1,
                    "servings": 1.0,
                    "servings_source": "inferred_single_person_meal",
                    "origin": "base",
                    "immutable": True,
                    "status": "active",
                    "composition_status": composition_status,
                    "ingredient_lines": lines,
                    "source_ingredient_text": str(meal.get("ingredients") or ""),
                    "nutrition": {
                        "mode": nutrition_mode,
                        "values_per_serving": public_nutrients(calculated if calculated is not None else source_values),
                        "source_values_per_serving": public_nutrients(source_values),
                        "rounded_match_to_source": nutrient_match,
                        "calculation_version": "v5-core-1",
                    },
                    "prep_minutes": int(meal["prep_minutes"]),
                    "meal_prep": {
                        "prepare_ahead": meal["prepare_ahead"],
                        "cold": meal["cold"],
                        "reheat": meal["reheat"],
                        "fridge": meal["fridge"],
                    },
                    "cuisine": meal["cuisine"],
                    "spices": meal["spices"],
                    "instructions": [],
                    "instructions_status": "missing",
                    "practical_notes": meal["notes"],
                    "editable_ingredient_composition": composition_status == "structured",
                    "source_occurrence_ids": [],
                }
            version_occurrences[version_id].append(occurrence_id)

            plan_day["meals"].append(
                {
                    "id": occurrence_id,
                    "sequence": sequence,
                    "recipe_id": family_id,
                    "recipe_version_id": version_id,
                    "time": meal["time"],
                    "minute_of_day": meal["minute_of_day"],
                    "day_offset": meal["day_offset"],
                    "meal_type": meal["meal_type"],
                    "source_anchor": meal["anchor"],
                    "flexible": bool(meal["flexible"]),
                    "status": "planned",
                }
            )
        plan_days.append(plan_day)

    if parse_errors:
        raise RuntimeError("Errori di parsing ingredienti:\n" + "\n".join(parse_errors[:30]))

    # Convert sets and attach occurrences deterministically.
    family_rows: list[dict[str, Any]] = []
    for family in sorted(recipe_families.values(), key=lambda item: item["title"].casefold()):
        version_ids = sorted(family["version_ids"])
        default_version_id = min(
            version_ids,
            key=lambda version_id: (-len(version_occurrences[version_id]), version_id),
        )
        family_rows.append(
            {
                **family,
                "meal_types": sorted(family["meal_types"], key=str.casefold),
                "cuisines": sorted(family["cuisines"], key=str.casefold),
                "version_ids": version_ids,
                "default_version_id": default_version_id,
            }
        )
    version_rows = []
    for version in sorted(recipe_versions.values(), key=lambda item: item["id"]):
        version["source_occurrence_ids"] = sorted(version_occurrences[version["id"]])
        version["source_occurrence_count"] = len(version["source_occurrence_ids"])
        version_rows.append(version)

    # Exact reconciliation with the workbook's daily ingredient detail.
    reconciliation_differences: list[dict[str, Any]] = []
    for day_number in range(1, 181):
        actual = {
            (item["code"], item["unit"]): float(item["quantity"])
            for item in ingredient_days.get(day_number, [])
        }
        parsed = dict(parsed_day_totals.get(day_number, {}))
        for key in sorted(set(actual) | set(parsed)):
            expected = actual.get(key, 0.0)
            observed = parsed.get(key, 0.0)
            if not math.isclose(expected, observed, abs_tol=0.01):
                reconciliation_differences.append(
                    {
                        "global_day": day_number,
                        "code": key[0],
                        "unit": key[1],
                        "workbook_quantity": expected,
                        "parsed_quantity": observed,
                        "difference": observed - expected,
                    }
                )

    # Usage and ingredient base snapshot.
    usage_days: dict[str, set[int]] = defaultdict(set)
    usage_total: dict[str, float] = defaultdict(float)
    for day_number, items in ingredient_days.items():
        for item in items:
            usage_days[item["code"]].add(day_number)
            usage_total[item["code"]] += float(item["quantity"])

    ingredient_rows: list[dict[str, Any]] = []
    ingredient_audit_rows: list[dict[str, Any]] = []
    state_counter: Counter[str] = Counter()
    for item in sorted(catalog_rows, key=lambda row: row["code"]):
        state, state_source, state_review_required = infer_state(item["name"])
        state_counter[state] += 1
        aliases = sorted(
            {
                label
                for label in alias_labels.get(item["code"], set())
                if normalize_text(label) != normalize_text(item["name"])
            },
            key=str.casefold,
        )
        conversions = []
        if item["code"] == "uovo":
            conversions.append(
                {
                    "id": "base:conversion:uovo-piece-50g",
                    "unit": "piece",
                    "singular_label": "uovo",
                    "plural_label": "uova",
                    "base_quantity": 50.0,
                    "base_unit": "g",
                    "source": "legacy_workbook_note",
                }
            )
        ingredient_record = {
            "id": f"base:ingredient:{item['code']}",
            "revision_id": f"base:ingredient-revision:{item['code']}@1",
            "code": item["code"],
            "name": item["name"],
            "aliases": aliases,
            "category_id": build_site.slugify(item["category"]),
            "category_name": item["category"],
            "brand": None,
            "origin": "base",
            "immutable": True,
            "status": "active",
            "revision": 1,
            "food_state": state,
            "food_state_source": state_source,
            "food_state_review_required": state_review_required,
            "nutrition_basis": {"amount": 100.0, "unit": item["unit"]},
            "nutrients": {
                "energy_kcal": item["kcal"],
                "protein_g": item["protein"],
                "carbohydrate_g": item["carbs"],
                "fat_g": item["fat"],
                "fiber_g": item["fiber"],
                "sugars_g": None,
                "saturated_fat_g": None,
                "salt_g": None,
                "sodium_mg": None,
            },
            "conversions": conversions,
            "provenance": {
                "kind": "legacy_dataset_estimate",
                "granularity": "dataset_level",
                "source_name": "Workbook TataDiet revisionato; riferimento generale CREA o etichetta del prodotto",
                "source_url": "https://www.alimentinutrizione.it/tabelle-nutrizionali/ricerca-per-alimento",
                "source_file": SOURCE_XLSX.name,
                "source_sheet": "Ingredienti",
                "source_row": item["source_row"],
                "captured_at": None,
                "note": "Il workbook non conserva una fonte puntuale per ogni singolo alimento.",
            },
            "usage": {
                "used_in_base_plan": item["code"] in usage_days,
                "occurrence_days": len(usage_days.get(item["code"], set())),
                "total_base_quantity": round(usage_total.get(item["code"], 0.0), 6),
                "base_unit": item["unit"],
            },
        }
        ingredient_rows.append(ingredient_record)
        estimated_kcal = 4 * item["protein"] + 4 * item["carbs"] + 9 * item["fat"]
        ingredient_audit_rows.append(
            {
                "code": item["code"],
                "name": item["name"],
                "category": item["category"],
                "base_unit": item["unit"],
                "kcal": item["kcal"],
                "protein_g": item["protein"],
                "carbohydrate_g": item["carbs"],
                "fat_g": item["fat"],
                "fiber_g": item["fiber"],
                "atwater_estimate_kcal": round(estimated_kcal, 2),
                "atwater_delta_kcal": round(item["kcal"] - estimated_kcal, 2),
                "used_in_plan": item["code"] in usage_days,
                "occurrence_days": len(usage_days.get(item["code"], set())),
                "food_state": state,
                "state_review_required": state_review_required,
                "aliases": " | ".join(aliases),
                "conversion_count": len(conversions),
                "provenance_granularity": "dataset_level",
                "optional_nutrients_present": 0,
                "audit_status": "review_state" if state_review_required else "core_ready",
            }
        )

    used_codes = set(usage_days)
    catalog_codes = set(catalog)
    multi_version_families = sum(1 for family in family_rows if len(family["version_ids"]) > 1)
    flexible_versions = sum(1 for version in version_rows if version["composition_status"] == "unstructured_estimate")

    base_generated_at = stable_source_timestamp(workbook)
    ingredients_payload = {
        "format": "tatadiet-base-ingredients",
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": base_generated_at,
        "ingredients": ingredient_rows,
    }
    recipes_payload = {
        "format": "tatadiet-base-recipes",
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": base_generated_at,
        "recipe_families": family_rows,
        "recipe_versions": version_rows,
    }
    plan_payload = {
        "format": "tatadiet-base-plan-template",
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": base_generated_at,
        "duration_days": 180,
        "start_mapping": {"cycle": 1, "variant": 1, "day_type": "D1", "base_global_day": 1},
        "summary": public_nutrients(overall_summary),
        "days": plan_days,
    }

    ingredients_path = BASE_DIR / "ingredients.base.v1.json"
    recipes_path = BASE_DIR / "recipes.base.v1.json"
    plan_path = BASE_DIR / "plan-template.base.v1.json"
    write_json(ingredients_path, ingredients_payload)
    write_json(recipes_path, recipes_payload)
    write_json(plan_path, plan_payload)

    source_hashes = {
        "xlsx": sha256_file(SOURCE_XLSX),
        "pdf": sha256_file(SOURCE_PDF),
    }
    manifest = {
        "format": "tatadiet-base-manifest",
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "phase_version": PHASE_VERSION,
        "generated_at": base_generated_at,
        "immutable": True,
        "source": {
            "xlsx": {"file": SOURCE_XLSX.name, "sha256": source_hashes["xlsx"]},
            "pdf": {"file": SOURCE_PDF.name, "sha256": source_hashes["pdf"]},
        },
        "files": {
            ingredients_path.name: {"sha256": sha256_file(ingredients_path), "records": len(ingredient_rows)},
            recipes_path.name: {
                "sha256": sha256_file(recipes_path),
                "recipe_families": len(family_rows),
                "recipe_versions": len(version_rows),
            },
            plan_path.name: {"sha256": sha256_file(plan_path), "days": len(plan_days), "meals": sum(len(day["meals"]) for day in plan_days)},
        },
        "invariants": [
            "Base records are immutable at runtime.",
            "User edits create user-owned records or occurrence overrides.",
            "The source workbook is never modified by the web application.",
            "A base dataset version is referenced by every full backup.",
        ],
    }
    write_json(BASE_DIR / "base-dataset-manifest.json", manifest)

    # CSV reports.
    ingredient_csv = AUDIT_DIR / "ingredient-audit.csv"
    with ingredient_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(ingredient_audit_rows[0].keys()))
        writer.writeheader()
        writer.writerows(ingredient_audit_rows)

    version_audit_rows = []
    for version in version_rows:
        source_values = version["nutrition"]["source_values_per_serving"]
        calculated = version["nutrition"]["values_per_serving"]
        version_audit_rows.append(
            {
                "recipe_id": version["recipe_id"],
                "version_id": version["id"],
                "composition_status": version["composition_status"],
                "ingredient_count": len(version["ingredient_lines"]),
                "occurrence_count": version["source_occurrence_count"],
                "source_kcal": source_values["energy_kcal"],
                "calculated_kcal": calculated["energy_kcal"],
                "rounded_match_to_source": version["nutrition"]["rounded_match_to_source"],
                "servings": version["servings"],
                "servings_source": version["servings_source"],
                "instructions_status": version["instructions_status"],
                "editable_ingredient_composition": version["editable_ingredient_composition"],
            }
        )
    version_csv = AUDIT_DIR / "recipe-version-audit.csv"
    with version_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(version_audit_rows[0].keys()))
        writer.writeheader()
        writer.writerows(version_audit_rows)

    dependency_map = scan_static_dependencies()
    write_json(AUDIT_DIR / "static-data-dependency-map.json", dependency_map)

    migration_map = {
        "format": "tatadiet-v4-to-v5-migration-map",
        "schema_version": 1,
        "generated_at": now_iso(),
        "source_version": "4.0.0",
        "target_phase": PHASE_VERSION,
        "local_storage": [
            {
                "source_key": "diet-plan:start-date:v2",
                "target_store": "settings",
                "target_key": "planStartDate",
                "strategy": "copy_if_valid_iso_date",
            },
            {
                "source_key_pattern": "diet-plan-shopping:*",
                "target_store": "shoppingChecklists",
                "strategy": "import_legacy_static_checklist",
            },
            {
                "source_key_pattern": "diet-plan-shopping-range:*",
                "target_store": "shoppingChecklists",
                "strategy": "import_legacy_range_checklist",
            },
        ],
        "legacy_export": {
            "format": "diet-plan-preferences",
            "version": 1,
            "strategy": "accept_settings_and_checklists_only",
        },
        "not_migrated": [
            "Cache API entries and service-worker caches",
            "Generated HTML and static JSON files",
            "URL query parameters other than a valid start date",
        ],
    }
    write_json(AUDIT_DIR / "v4-to-v5-migration-map.json", migration_map)

    formula_audit = audit_workbook_formulas(SOURCE_XLSX)
    nutrient_max_absolute_difference = {
        key: round(max(values), 6) if values else None
        for key, values in nutrient_absolute_differences.items()
    }

    known_gaps = [
        {
            "id": "ingredient-state",
            "severity": "medium",
            "count": state_counter["unspecified"],
            "description": "Lo stato crudo/cotto/secco/sgocciolato non è dichiarato esplicitamente per tutti gli ingredienti.",
            "phase_action": "Conservare 'unspecified' nel base dataset; richiederlo per i nuovi ingredienti personali quando rilevante.",
        },
        {
            "id": "item-provenance",
            "severity": "medium",
            "count": len(ingredient_rows),
            "description": "La fonte nutrizionale è documentata a livello di dataset, non per singolo ingrediente.",
            "phase_action": "Per gli ingredienti personali registrare fonte, etichetta o nota manuale a livello record.",
        },
        {
            "id": "optional-nutrients",
            "severity": "low",
            "count": len(ingredient_rows),
            "description": "Zuccheri, saturi, sale e sodio non sono presenti nel workbook.",
            "phase_action": "Campi previsti ma facoltativi nello schema V5; non bloccano il calcolo core.",
        },
        {
            "id": "recipe-instructions",
            "severity": "medium",
            "count": len(family_rows),
            "description": "Le ricette base non contengono passaggi di preparazione strutturati.",
            "phase_action": "Mantenere instructions_status='missing'; consentire istruzioni nelle ricette personali.",
        },
        {
            "id": "implicit-servings",
            "severity": "low",
            "count": len(version_rows),
            "description": "Il numero di porzioni non è esplicito; ogni voce è trattata come una porzione individuale.",
            "phase_action": "Base servings=1 con provenienza 'inferred'; obbligatorio nelle nuove ricette.",
        },
        {
            "id": "flexible-unstructured",
            "severity": "medium",
            "count": flexible_versions,
            "description": "I pasti flessibili hanno nutrienti stimati ma non una composizione ingredienti strutturata.",
            "phase_action": "Non ricalcolarli automaticamente; per modificarli occorre duplicare e specificare ingredienti oppure usare stima manuale.",
        },
    ]

    summary = {
        "format": "tatadiet-v5-phase1-audit",
        "schema_version": 1,
        "phase_version": PHASE_VERSION,
        "dataset_version": DATASET_VERSION,
        "generated_at": now_iso(),
        "status": "pass_with_documented_gaps",
        "source": {
            "xlsx": {"file": SOURCE_XLSX.name, "sha256": source_hashes["xlsx"]},
            "pdf": {"file": SOURCE_PDF.name, "sha256": source_hashes["pdf"]},
            "workbook": {
                "sheet_count": len(workbook.sheetnames),
                "sheets": list(workbook.sheetnames),
                "source_timestamp": stable_source_timestamp(workbook),
            },
        },
        "counts": {
            "ingredient_catalog": len(catalog),
            "used_ingredients": len(used_codes),
            "unused_ingredients": len(catalog_codes - used_codes),
            "ingredient_detail_rows": workbook["Dettaglio ingredienti"].max_row - 1,
            "ingredient_day_records": sum(len(items) for items in ingredient_days.values()),
            "days": len(days),
            "meals": sum(len(day["meals"]) for day in days),
            "structured_meals": structured_meals,
            "estimated_flexible_meals": estimated_meals,
            "meal_ingredient_lines": ingredient_line_count,
            "recipe_families": len(family_rows),
            "recipe_versions": len(version_rows),
            "families_with_multiple_versions": multi_version_families,
            "flexible_recipe_versions": flexible_versions,
            "explicit_or_inferred_state": len(ingredient_rows) - state_counter["unspecified"],
            "state_unspecified": state_counter["unspecified"],
        },
        "validation": {
            "catalog_missing_core_values": sum(
                1 for item in catalog_rows if any(item[key] is None for key in ("kcal", "protein", "carbs", "fat", "fiber"))
            ),
            "used_codes_missing_from_catalog": sorted(used_codes - catalog_codes),
            "catalog_codes_unused_in_plan": sorted(catalog_codes - used_codes),
            "meal_parse_errors": len(parse_errors),
            "daily_reconciliation_differences": len(reconciliation_differences),
            "structured_nutrient_matches_at_0_1": structured_nutrient_matches,
            "structured_nutrient_mismatches_at_0_1": structured_meals - structured_nutrient_matches,
            "duplicate_ingredient_codes": [code for code, count in Counter(item["code"] for item in catalog_rows).items() if count > 1],
            "duplicate_ingredient_names": [
                name for name, count in Counter(normalize_text(item["name"]) for item in catalog_rows).items() if count > 1
            ],
            "nutrient_max_absolute_difference": nutrient_max_absolute_difference,
            "workbook_formulas": formula_audit,
        },
        "state_breakdown": dict(sorted(state_counter.items())),
        "known_gaps": known_gaps,
        "release_gate": {
            "phase_1_complete": (
                len(days) == 180
                and ingredient_line_count == 3189
                and not parse_errors
                and not reconciliation_differences
                and structured_nutrient_matches == structured_meals
                and len(catalog) == 130
                and len(family_rows) == 306
                and len(version_rows) == 547
                and formula_audit["formula_count"] == 5308
                and formula_audit["missing_cached_values"] == 0
                and not formula_audit["cached_errors"]
                and all(
                    value is not None and value <= 0.051
                    for value in nutrient_max_absolute_difference.values()
                )
            ),
            "phase_2_prerequisites": [
                "Implementare IndexedDB usando gli schemi congelati.",
                "Importare la data iniziale e le checklist V4 secondo la mappa di migrazione.",
                "Non rendere mutabili i record base.",
            ],
        },
    }
    write_json(AUDIT_DIR / "audit-summary.json", summary)
    write_json(AUDIT_DIR / "daily-reconciliation-differences.json", reconciliation_differences)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if not summary["release_gate"]["phase_1_complete"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
