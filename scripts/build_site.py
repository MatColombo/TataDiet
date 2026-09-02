#!/usr/bin/env python3
"""Build the static diet-plan website from the reviewed Excel workbook."""

from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import unicodedata
from collections import Counter, OrderedDict, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from typing import Any, Iterable
from urllib.parse import quote

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
STATIC = ROOT / "static"
TEMPLATES = ROOT / "templates"
SOURCE_XLSX = ROOT / "source_data" / "Piano_alimentare_revisionato_6_mesi_fibra_moderata.xlsx"
SOURCE_PDF = ROOT / "source_data" / "Piano_alimentare_revisionato_6_mesi_fibra_moderata.pdf"
DOWNLOAD_XLSX = DOCS / "downloads" / SOURCE_XLSX.name
DOWNLOAD_PDF = DOCS / "downloads" / "Piano_alimentare_revisionato_6_mesi_fibra_moderata.pdf"
VERSION = "5.2.1"

MONTH_SHEETS = [
    (1, "M1 Settembre"),
    (2, "M2 Ottobre"),
    (3, "M3 Novembre"),
    (4, "M4 Dicembre"),
    (5, "M5 Gennaio"),
    (6, "M6 Febbraio"),
]

SHIFT_INFO = {
    "D1": {"name": "Giornata", "short": "G", "hours": "08:00–20:00"},
    "D2": {"name": "Notte", "short": "N", "hours": "20:00–08:00; ultimo mini-pasto alle 08:20"},
    "D3": {"name": "Smonto", "short": "SN", "hours": "Ripresa alimentare dal risveglio pomeridiano"},
    "D4": {"name": "Riposo 1", "short": "R1", "hours": "Giornata di riposo"},
    "D5": {"name": "Riposo 2", "short": "R2", "hours": "Giornata di riposo"},
    "M": {"name": "Mattino", "short": "M", "hours": "Profilo alimentare di Giornata"},
    "P": {"name": "Pomeriggio", "short": "P", "hours": "Profilo alimentare di Giornata"},
}
DAY_UI = {code: data for code, data in SHIFT_INFO.items()}
DAY_UI.update({"CUSTOM":{"name":"Personalizzata","short":"C"},"OFF":{"name":"Fuori servizio","short":"OFF"},"FREE":{"name":"Giornata libera","short":"L"}})
def day_short(code: str) -> str:
    return DAY_UI.get(code, {"short": code}).get("short", code)
def day_label(code: str) -> str:
    return DAY_UI.get(code, {"name": code}).get("name", code)

CATEGORY_ORDER = [
    "Ortofrutta",
    "Carne e affettati",
    "Pesce",
    "Latticini e uova",
    "Cereali, pane e derivati",
    "Legumi e conserve",
    "Frutta secca e semi",
    "Condimenti e dispensa",
    "Pasto flessibile",
]


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value or "elemento"


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def yes(value: Any) -> bool:
    return str(value or "").strip().lower() in {"sì", "si", "yes", "true", "1"}


def avg(items: Iterable[dict[str, Any]], key: str) -> float:
    values = [as_float(item.get(key)) for item in items]
    return mean(values) if values else 0.0


def format_quantity(value: Any) -> str:
    number = as_float(value)
    if math.isclose(number, round(number), abs_tol=1e-9):
        return f"{int(round(number)):,}".replace(",", ".")
    return f"{number:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")



def positive_option(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text or text.startswith("no") or text.startswith("non ") or text.startswith("meglio al momento"):
        return False
    return any(token in text for token in ("sì", "si", "buono", "accettabile", "facoltativo", "ottimo", "giorno prima", "sera prima", "batch"))


def parse_fridge_days(value: Any) -> float:
    text = str(value or "").strip().lower()
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*giorn", text)
    if match:
        return float(match.group(1).replace(",", "."))
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*or", text)
    if match:
        return float(match.group(1).replace(",", ".")) / 24.0
    return 0.0


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean_docs() -> None:
    """Recreate docs from canonical static assets before rendering pages."""
    if DOCS.exists():
        shutil.rmtree(DOCS)
    DOCS.mkdir(parents=True, exist_ok=True)
    if not STATIC.exists():
        raise FileNotFoundError(f"Cartella static mancante: {STATIC}")
    shutil.copytree(STATIC, DOCS, dirs_exist_ok=True)
    (DOCS / "data").mkdir(parents=True, exist_ok=True)
    (DOCS / "downloads").mkdir(parents=True, exist_ok=True)
    (DOCS / ".nojekyll").touch()


def build_offline_manifest() -> dict[str, Any]:
    """Describe the site resources available for the optional full offline pack."""
    allowed_suffixes = {".html", ".json", ".css", ".js", ".svg", ".png", ".webmanifest"}
    excluded = {"service-worker.js", "data/offline-assets.json"}
    assets: list[str] = []
    total_bytes = 0
    for path in sorted(DOCS.rglob("*")):
        if not path.is_file():
            continue
        relative = path.relative_to(DOCS).as_posix()
        if relative in excluded or relative.startswith("downloads/") or path.suffix.lower() not in allowed_suffixes:
            continue
        assets.append(relative)
        total_bytes += path.stat().st_size
    payload = {
        "version": VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "strategy": "full-static-library",
        "asset_count": len(assets),
        "total_bytes": total_bytes,
        "assets": assets,
    }
    write_json(DOCS / "data" / "offline-assets.json", payload)
    return payload


def parse_summary(workbook) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    ws = workbook["Sintesi"]
    monthly = []
    for row in range(4, 10):
        monthly.append(
            {
                "month": ws.cell(row, 1).value,
                "kcal": as_float(ws.cell(row, 2).value),
                "protein": as_float(ws.cell(row, 3).value),
                "carbs": as_float(ws.cell(row, 4).value),
                "fat": as_float(ws.cell(row, 5).value),
                "fiber": as_float(ws.cell(row, 6).value),
                "days": int(as_float(ws.cell(row, 7).value)),
                "flexible": int(as_float(ws.cell(row, 8).value)),
            }
        )
    overall = {
        "kcal": as_float(ws.cell(11, 2).value),
        "protein": as_float(ws.cell(11, 3).value),
        "carbs": as_float(ws.cell(11, 4).value),
        "fat": as_float(ws.cell(11, 5).value),
        "fiber": as_float(ws.cell(11, 6).value),
        "days": int(as_float(ws.cell(11, 7).value)),
        "flexible": int(as_float(ws.cell(11, 8).value)),
    }
    return monthly, overall


def parse_plan(workbook, monthly_summary: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    days_by_number: dict[int, dict[str, Any]] = {}
    season_notes: dict[int, str] = {}

    for cycle_number, sheet_name in MONTH_SHEETS:
        ws = workbook[sheet_name]
        season_notes[cycle_number] = str(ws.cell(2, 1).value or "")
        last_day: dict[str, Any] | None = None
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
            first = row[0]
            if isinstance(first, (int, float)) and row[1] is not None:
                global_day = int(first)
                variant_number = int(row[1])
                type_text = str(row[2] or "")
                d_code = type_text.split(" - ", 1)[0].strip()
                shift_name = type_text.split(" - ", 1)[1].strip() if " - " in type_text else SHIFT_INFO.get(d_code, {}).get("name", type_text)
                day = days_by_number.setdefault(
                    global_day,
                    {
                        "global_day": global_day,
                        "cycle": cycle_number,
                        "month": str(row[21] or monthly_summary[cycle_number - 1]["month"]),
                        "variant": variant_number,
                        "day_in_variant": ((global_day - 1) % 5) + 1,
                        "d_code": d_code,
                        "shift_name": shift_name,
                        "shift_hours": SHIFT_INFO.get(d_code, {}).get("hours", ""),
                        "meals": [],
                        "total": {},
                        "flexible": False,
                    },
                )
                title = str(row[5] or "Pasto senza titolo").strip()
                time_text = str(row[3] or "").strip()
                anchor = f"pasto-{time_text.replace(':', '').replace('.', '') or len(day['meals']) + 1}"
                try:
                    hour, minute = (int(part) for part in time_text.split(":"))
                    minute_of_day = hour * 60 + minute
                except (TypeError, ValueError):
                    minute_of_day = len(day["meals"]) * 60
                day_offset = 0
                if day["meals"]:
                    previous_absolute = day["meals"][-1]["minute_of_day"] + day["meals"][-1]["day_offset"] * 1440
                    while minute_of_day + day_offset * 1440 <= previous_absolute:
                        day_offset += 1
                meal = {
                    "time": time_text,
                    "minute_of_day": minute_of_day,
                    "day_offset": day_offset,
                    "meal_type": str(row[4] or "Pasto").strip(),
                    "title": title,
                    "ingredients": str(row[6] or "").strip(),
                    "prep_minutes": int(round(as_float(row[7]))),
                    "kcal": as_float(row[8]),
                    "protein": as_float(row[9]),
                    "carbs": as_float(row[10]),
                    "fat": as_float(row[11]),
                    "fiber": as_float(row[12]),
                    "cuisine": str(row[13] or "Non specificata").strip(),
                    "spices": str(row[14] or "Nessuna").strip(),
                    "prepare_ahead": str(row[15] or "Non specificato").strip(),
                    "cold": str(row[16] or "Non specificato").strip(),
                    "reheat": str(row[17] or "Non specificato").strip(),
                    "fridge": str(row[18] or "Non specificato").strip(),
                    "notes": str(row[19] or "").strip(),
                    "flexible": yes(row[20]),
                    "anchor": anchor,
                }
                day["meals"].append(meal)
                day["flexible"] = day["flexible"] or meal["flexible"]
                last_day = day
            elif isinstance(first, str) and first.startswith("Totale D") and last_day is not None:
                last_day["total"] = {
                    "kcal": as_float(row[8]),
                    "protein": as_float(row[9]),
                    "carbs": as_float(row[10]),
                    "fat": as_float(row[11]),
                    "fiber": as_float(row[12]),
                }
                last_day["flexible"] = last_day["flexible"] or yes(row[20])

    days = [days_by_number[index] for index in sorted(days_by_number)]
    for day in days:
        if not day["total"]:
            day["total"] = {key: sum(as_float(meal[key]) for meal in day["meals"]) for key in ("kcal", "protein", "carbs", "fat", "fiber")}

    title_to_slug: dict[str, str] = {}
    used_slugs: dict[str, str] = {}
    for title in sorted({meal["title"] for day in days for meal in day["meals"]}, key=str.casefold):
        base = slugify(title)
        slug = base
        sequence = 2
        while slug in used_slugs and used_slugs[slug] != title:
            slug = f"{base}-{sequence}"
            sequence += 1
        used_slugs[slug] = title
        title_to_slug[title] = slug

    recipes_raw: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for day in days:
        for meal in day["meals"]:
            meal["recipe_slug"] = title_to_slug[meal["title"]]
            occurrence = {
                **meal,
                "cycle": day["cycle"],
                "month": day["month"],
                "variant": day["variant"],
                "global_day": day["global_day"],
                "d_code": day["d_code"],
            }
            recipes_raw[meal["title"]].append(occurrence)

    recipes: dict[str, dict[str, Any]] = {}
    for title, occurrences in recipes_raw.items():
        meal_types = sorted({item["meal_type"] for item in occurrences}, key=str.casefold)
        cuisines = sorted({item["cuisine"] for item in occurrences}, key=str.casefold)
        version_map: OrderedDict[tuple[Any, ...], dict[str, Any]] = OrderedDict()
        for item in occurrences:
            key = (
                item["ingredients"], round(item["kcal"], 1), item["prep_minutes"], item["prepare_ahead"],
                item["cold"], item["reheat"], item["fridge"], item["cuisine"], item["spices"],
            )
            if key not in version_map:
                version_map[key] = {
                    "ingredients": item["ingredients"],
                    "kcal": item["kcal"],
                    "protein": item["protein"],
                    "carbs": item["carbs"],
                    "fat": item["fat"],
                    "fiber": item["fiber"],
                    "prep_minutes": item["prep_minutes"],
                    "prepare_ahead": item["prepare_ahead"],
                    "cold": item["cold"],
                    "reheat": item["reheat"],
                    "fridge": item["fridge"],
                    "cuisine": item["cuisine"],
                    "spices": item["spices"],
                }
        primary_meal = Counter(item["meal_type"] for item in occurrences).most_common(1)[0][0]
        primary_cuisine = Counter(item["cuisine"] for item in occurrences).most_common(1)[0][0]
        ahead = any(
            any(token in item["prepare_ahead"].lower() for token in ("sì", "si,", "prima", "anticipo", "batch", "sera"))
            and not item["prepare_ahead"].lower().startswith("no")
            for item in occurrences
        )
        sample_ingredients = next((item["ingredients"] for item in occurrences if item["ingredients"]), "Pasto flessibile")
        slug = title_to_slug[title]
        recipe = {
            "title": title,
            "slug": slug,
            "meal_types": meal_types,
            "meal_types_lower": " ".join(item.lower() for item in meal_types),
            "cuisines": cuisines,
            "cuisines_lower": " ".join(item.lower() for item in cuisines),
            "primary_meal": primary_meal,
            "primary_cuisine": primary_cuisine,
            "occurrence_count": len(occurrences),
            "avg_kcal": avg(occurrences, "kcal"),
            "avg_protein": avg(occurrences, "protein"),
            "avg_carbs": avg(occurrences, "carbs"),
            "avg_fat": avg(occurrences, "fat"),
            "avg_fiber": avg(occurrences, "fiber"),
            "avg_prep": avg(occurrences, "prep_minutes"),
            "min_prep": min(item["prep_minutes"] for item in occurrences),
            "max_prep": max(item["prep_minutes"] for item in occurrences),
            "ahead": ahead,
            "cold_ok": any(positive_option(item["cold"]) for item in occurrences),
            "reheat_ok": any(positive_option(item["reheat"]) for item in occurrences),
            "max_fridge_days": max(parse_fridge_days(item["fridge"]) for item in occurrences),
            "sample_ingredients": sample_ingredients,
            "search_text": " ".join([title, *sorted({item["ingredients"] for item in occurrences if item["ingredients"]}), *meal_types, *cuisines]).lower(),
            "versions": list(version_map.values()),
            "occurrences": [
                {
                    "cycle": item["cycle"], "month": item["month"], "variant": item["variant"],
                    "global_day": item["global_day"], "d_code": item["d_code"], "time": item["time"],
                    "meal_type": item["meal_type"], "anchor": item["anchor"],
                }
                for item in occurrences
            ],
        }
        recipes[slug] = recipe

    cycles: list[dict[str, Any]] = []
    for cycle_number, _sheet_name in MONTH_SHEETS:
        cycle_days = [day for day in days if day["cycle"] == cycle_number]
        variants = []
        for variant_number in range(1, 7):
            variant_days = [day for day in cycle_days if day["variant"] == variant_number]
            variants.append(
                {
                    "number": variant_number,
                    "start_day": min(day["global_day"] for day in variant_days),
                    "end_day": max(day["global_day"] for day in variant_days),
                    "days": variant_days,
                    "flexible_count": sum(1 for day in variant_days if day["flexible"]),
                    "avg": {key: mean(day["total"][key] for day in variant_days) for key in ("kcal", "protein", "carbs", "fat", "fiber")},
                }
            )
        cycles.append(
            {
                "number": cycle_number,
                "month": monthly_summary[cycle_number - 1]["month"],
                "season_note": season_notes[cycle_number],
                "avg": {key: monthly_summary[cycle_number - 1][key] for key in ("kcal", "protein", "carbs", "fat", "fiber")},
                "flexible_count": monthly_summary[cycle_number - 1]["flexible"],
                "variants": variants,
            }
        )

    return cycles, days, recipes


def parse_shopping(workbook) -> tuple[dict[tuple[int, int], list[dict[str, Any]]], dict[int, list[dict[str, Any]]]]:
    variant_groups: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    ws = workbook["Spesa per variante"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not isinstance(row[0], (int, float)) or row[2] is None:
            continue
        item = {
            "category": str(row[3] or "Altro"),
            "code": str(row[4] or slugify(str(row[5] or "elemento"))),
            "name": str(row[5] or "Elemento"),
            "exact": as_float(row[6]),
            "unit": str(row[7] or ""),
            "rounded": as_float(row[8]),
            "notes": str(row[9] or ""),
        }
        # Nel foglio di origine le uova hanno il fabbisogno esatto espresso in grammi
        # ma l'unità di acquisto è "uova". Normalizziamo a numero di uova (50 g cad.).
        if item["code"] == "uovo" and item["unit"].casefold() == "uova":
            item["exact"] = item["exact"] / 50.0
        variant_groups[(int(row[0]), int(row[2]))].append(item)

    cycle_groups: dict[int, list[dict[str, Any]]] = defaultdict(list)
    ws = workbook["Spesa per ciclo"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not isinstance(row[0], (int, float)) or row[2] is None:
            continue
        item = {
            "category": str(row[2] or "Altro"),
            "code": str(row[3] or slugify(str(row[4] or "elemento"))),
            "name": str(row[4] or "Elemento"),
            "exact": as_float(row[5]),
            "unit": str(row[6] or ""),
            "rounded": as_float(row[7]),
            "notes": str(row[8] or ""),
        }
        if item["code"] == "uovo" and item["unit"].casefold() == "uova":
            item["exact"] = item["exact"] / 50.0
        cycle_groups[int(row[0])].append(item)

    def sort_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        order = {name: index for index, name in enumerate(CATEGORY_ORDER)}
        return sorted(items, key=lambda item: (order.get(item["category"], 999), item["category"].casefold(), item["name"].casefold()))

    return ({key: sort_items(value) for key, value in variant_groups.items()}, {key: sort_items(value) for key, value in cycle_groups.items()})


def parse_ingredient_details(workbook) -> tuple[dict[int, list[dict[str, Any]]], dict[str, dict[str, Any]]]:
    ws = workbook["Dettaglio ingredienti"]
    day_groups: dict[int, dict[tuple[str, str], dict[str, Any]]] = defaultdict(dict)
    catalog: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not isinstance(row[3], (int, float)) or row[5] is None:
            continue
        global_day = int(row[3])
        code = str(row[5]).strip()
        name = str(row[6] or code).strip()
        category = str(row[7] or "Altro").strip()
        quantity = as_float(row[8])
        unit = str(row[9] or "").strip()
        key = (code, unit)
        existing = day_groups[global_day].get(key)
        if existing:
            existing["quantity"] += quantity
        else:
            day_groups[global_day][key] = {
                "code": code,
                "name": name,
                "category": category,
                "quantity": quantity,
                "unit": unit,
            }
        entry = catalog.setdefault(
            code,
            {
                "code": code,
                "name": name,
                "category": category,
                "unit": unit,
                "days": set(),
                "total": 0.0,
            },
        )
        entry["days"].add(global_day)
        entry["total"] += quantity

    order = {name: index for index, name in enumerate(CATEGORY_ORDER)}
    public_days: dict[int, list[dict[str, Any]]] = {}
    for global_day, items in day_groups.items():
        public_days[global_day] = sorted(
            items.values(),
            key=lambda item: (
                order.get(item["category"], 999),
                item["category"].casefold(),
                item["name"].casefold(),
            ),
        )
    public_catalog: dict[str, dict[str, Any]] = {}
    for code, item in catalog.items():
        public_catalog[code] = {
            **item,
            "days": sorted(item["days"]),
            "occurrence_days": len(item["days"]),
        }
    return public_days, public_catalog


def infer_rounding_rules(
    shopping_variants: dict[tuple[int, int], list[dict[str, Any]]],
    shopping_cycles: dict[int, list[dict[str, Any]]],
    catalog: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    observations: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for groups in (shopping_variants.values(), shopping_cycles.values()):
        for items in groups:
            for item in items:
                if item["exact"] > 0 and item["rounded"] > 0:
                    observations[(item["code"], item["unit"])].append(item)

    common_steps = [1, 2, 5, 10, 20, 25, 30, 40, 50, 75, 100, 125, 150, 200, 250, 300, 400, 500, 750, 1000]
    rules: dict[str, dict[str, Any]] = {}
    for code, base in catalog.items():
        unit = base["unit"]
        if code == "uovo":
            egg_rows = [item for items in shopping_variants.values() for item in items if item["code"] == "uovo"]
            notes = Counter(row["notes"] for row in egg_rows if row.get("notes"))
            rules[code] = {
                "code": code,
                "name": base["name"],
                "category": base["category"],
                "unit": unit,
                "display_unit": "uova",
                "conversion_factor": 50,
                "rounding_step": 1,
                "note": notes.most_common(1)[0][0] if notes else "1 uovo equivale indicativamente a 50 g.",
                "observations": len(egg_rows),
            }
            continue
        rows = observations.get((code, unit), [])
        candidates = [1] if unit.casefold() in {"unità", "unita", "porzione", "porzioni"} else common_steps
        valid: list[float] = []
        for step in candidates:
            matches = True
            for row in rows:
                expected = math.ceil((row["exact"] - 1e-9) / step) * step
                if not math.isclose(expected, row["rounded"], abs_tol=0.11):
                    matches = False
                    break
            if matches:
                valid.append(step)
        step = max(valid) if valid else (1 if unit.casefold() in {"unità", "unita", "porzione", "porzioni"} else 10)
        notes = Counter(row["notes"] for row in rows if row.get("notes"))
        rules[code] = {
            "code": code,
            "name": base["name"],
            "category": base["category"],
            "unit": unit,
            "rounding_step": step,
            "note": notes.most_common(1)[0][0] if notes else "Quantità suggerita arrotondata per eccesso.",
            "observations": len(rows),
        }
    return rules


def build_search_index(
    cycles: list[dict[str, Any]],
    days: list[dict[str, Any]],
    recipes: list[dict[str, Any]],
    ingredient_catalog: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for recipe in recipes:
        entries.append(
            {
                "id": f"recipe:{recipe['slug']}",
                "type": "recipe",
                "type_label": "Ricetta",
                "title": recipe["title"],
                "subtitle": f"{recipe['primary_meal']} · {recipe['primary_cuisine']} · {recipe['avg_prep']:.0f} min",
                "text": recipe["search_text"],
                "href": f"ricette/{recipe['slug']}/index.html",
                "badges": [f"{recipe['avg_kcal']:.0f} kcal", f"{recipe['occurrence_count']} occorrenze"],
            }
        )
    for ingredient in sorted(ingredient_catalog.values(), key=lambda item: item["name"].casefold()):
        entries.append(
            {
                "id": f"ingredient:{ingredient['code']}",
                "type": "ingredient",
                "type_label": "Ingrediente",
                "title": ingredient["name"],
                "subtitle": ingredient["category"],
                "text": f"{ingredient['name']} {ingredient['category']} {ingredient['code']}",
                "href": f"ricette/index.html?q={quote(ingredient['name'])}",
                "badges": [f"{ingredient['occurrence_days']} giorni", ingredient["unit"]],
            }
        )
    for day in days:
        meal_titles = " ".join(meal["title"] for meal in day["meals"])
        ingredients = " ".join(meal["ingredients"] for meal in day["meals"] if meal["ingredients"])
        entries.append(
            {
                "id": f"day:{day['global_day']}",
                "type": "day",
                "type_label": "Giorno",
                "title": f"Giorno {day['global_day']} · {day_short(day['d_code'])} · {day_label(day['d_code'])}",
                "subtitle": f"{day['month']} · C{day['cycle']} · V{day['variant']}",
                "text": f"{meal_titles} {ingredients} {day['month']} {day['shift_name']} {day['d_code']} {day_short(day['d_code'])} {day_label(day['d_code'])}",
                "href": f"piano/ciclo-{day['cycle']}/variante-{day['variant']}/{day['d_code'].lower()}/index.html",
                "badges": [f"{day['total']['kcal']:.0f} kcal", f"{len(day['meals'])} pasti"],
                "global_day": day["global_day"],
            }
        )
    for cycle in cycles:
        entries.append(
            {
                "id": f"cycle:{cycle['number']}",
                "type": "cycle",
                "type_label": "Ciclo",
                "title": f"Ciclo {cycle['number']} · {cycle['month']}",
                "subtitle": "30 giorni · 6 varianti",
                "text": f"{cycle['month']} {cycle['season_note']}",
                "href": f"piano/ciclo-{cycle['number']}/index.html",
                "badges": [f"{cycle['avg']['kcal']:.0f} kcal/die", f"{cycle['flexible_count']} flessibili"],
            }
        )
        for variant in cycle["variants"]:
            titles = " ".join(meal["title"] for day in variant["days"] for meal in day["meals"])
            entries.append(
                {
                    "id": f"variant:{cycle['number']}:{variant['number']}",
                    "type": "variant",
                    "type_label": "Variante",
                    "title": f"{cycle['month']} · Variante {variant['number']}",
                    "subtitle": f"Ciclo {cycle['number']} · giorni {variant['start_day']}-{variant['end_day']}",
                    "text": f"{cycle['month']} variante {variant['number']} {titles}",
                    "href": f"piano/ciclo-{cycle['number']}/variante-{variant['number']}/index.html",
                    "badges": ["5 giorni", f"{variant['avg']['kcal']:.0f} kcal/die"],
                }
            )
    return entries

def group_by_category(items: list[dict[str, Any]]) -> OrderedDict[str, list[dict[str, Any]]]:
    grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for item in items:
        grouped.setdefault(item["category"], []).append(item)
    return grouped


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(SOURCE_XLSX)

    clean_docs()
    shutil.copy2(SOURCE_XLSX, DOWNLOAD_XLSX)
    if SOURCE_PDF.exists():
        shutil.copy2(SOURCE_PDF, DOWNLOAD_PDF)

    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    monthly_summary, overall_summary = parse_summary(workbook)
    cycles, days, recipes = parse_plan(workbook, monthly_summary)
    shopping_variants, shopping_cycles = parse_shopping(workbook)
    ingredient_days, ingredient_catalog = parse_ingredient_details(workbook)
    rounding_rules = infer_rounding_rules(shopping_variants, shopping_cycles, ingredient_catalog)

    if len(cycles) != 6 or len(days) != 180 or sum(len(cycle["variants"]) for cycle in cycles) != 36:
        raise RuntimeError("Struttura del piano inattesa: attesi 6 cicli, 36 varianti e 180 giorni")
    if any(len(day["meals"]) == 0 for day in days):
        raise RuntimeError("Almeno un giorno non contiene pasti")

    env = Environment(
        loader=FileSystemLoader(TEMPLATES),
        autoescape=select_autoescape(["html", "xml"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )
    env.filters["format_quantity"] = format_quantity
    env.globals["site_version"] = VERSION
    env.globals["day_short"] = day_short
    env.globals["day_label"] = day_label

    def render(template_name: str, relative_path: str, **context: Any) -> None:
        output = DOCS / relative_path
        output.parent.mkdir(parents=True, exist_ok=True)
        relroot_raw = os.path.relpath(DOCS, output.parent).replace(os.sep, "/")
        relroot = "" if relroot_raw == "." else relroot_raw.rstrip("/") + "/"
        breadcrumbs = []
        for crumb in context.pop("breadcrumbs", []):
            href = crumb.get("path")
            breadcrumbs.append({"label": crumb["label"], "href": f"{relroot}{href}" if href else None})
        html = env.get_template(template_name).render(relroot=relroot, breadcrumbs=breadcrumbs, **context)
        output.write_text(html, encoding="utf-8")

    counts = {
        "cycles": len(cycles),
        "variants": sum(len(cycle["variants"]) for cycle in cycles),
        "days": len(days),
        "meals": sum(len(day["meals"]) for day in days),
        "recipes": len(recipes),
    }
    shifts = [{"code": code, **data} for code, data in SHIFT_INFO.items()]

    render(
        "home.html", "index.html", title="Panoramica", nav="home", page_id="home",
        cycles=cycles, counts=counts, summary=overall_summary, shifts=shifts,
    )
    render(
        "today.html", "oggi/index.html", title="Oggi", nav="today", page_id="today",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Oggi"}],
    )
    render(
        "calendar.html", "calendario/index.html", title="Calendario", nav="calendar", page_id="calendar",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Calendario"}],
    )
    render(
        "plan_editor.html", "calendario/modifica/index.html", title="Planner personale", nav="calendar", page_id="plan-editor",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Calendario", "path": "calendario/index.html"}, {"label": "Modifica piano"}],
    )
    render(
        "day_composer.html", "calendario/componi/index.html", title="Compositore della giornata", nav="calendar", page_id="day-composer",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Calendario", "path": "calendario/index.html"}, {"label": "Compositore"}],
    )
    render(
        "day_manager.html", "calendario/gestisci/index.html", title="Gestisci giornata", nav="calendar", page_id="day-manager",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Calendario", "path": "calendario/index.html"}, {"label": "Gestisci giornata"}],
    )
    render(
        "preferences.html", "preferenze/index.html", title="Preferenze alimentari", nav="preferences", page_id="preferences",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Preferenze alimentari"}],
    )
    render(
        "prep.html", "preparazioni/index.html", title="Preparazioni 48 ore", nav="tools", page_id="prep",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Preparazioni 48 ore"}],
    )
    render(
        "search.html", "cerca/index.html", title="Cerca", nav="search", page_id="search",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Cerca"}],
    )
    render(
        "tools.html", "strumenti/index.html", title="Utilità", nav="tools", page_id="tools",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Utilità"}],
    )
    render(
        "ingredients.html", "ingredienti/index.html", title="Ingredienti", nav="ingredients", page_id="ingredients",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Ingredienti"}],
    )
    render(
        "recipe_studio.html", "ricette/studio/index.html", title="Studio ricette", nav="recipes", page_id="recipe-studio",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Ricette", "path": "ricette/index.html"}, {"label": "Studio ricette"}],
    )
    render(
        "recipe_scheduler.html", "ricette/programma/index.html", title="Programma ricetta", nav="recipes", page_id="recipe-scheduler",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Ricette", "path": "ricette/index.html"}, {"label": "Programma nel calendario"}],
    )
    render(
        "plan_index.html", "piano/index.html", title="Piano", nav="plan", page_id="plan",
        cycles=cycles,
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Piano"}],
    )

    day_lookup = {day["global_day"]: day for day in days}
    for cycle in cycles:
        render(
            "cycle.html", f"piano/ciclo-{cycle['number']}/index.html", title=f"Ciclo {cycle['number']} · {cycle['month']}",
            nav="plan", page_id="cycle", cycle=cycle,
            breadcrumbs=[
                {"label": "Panoramica", "path": "index.html"},
                {"label": "Piano", "path": "piano/index.html"},
                {"label": f"Ciclo {cycle['number']} · {cycle['month']}"},
            ],
        )
        for variant_index, variant in enumerate(cycle["variants"]):
            prev_variant = cycle["variants"][variant_index - 1] if variant_index > 0 else None
            next_variant = cycle["variants"][variant_index + 1] if variant_index + 1 < len(cycle["variants"]) else None
            render(
                "variant.html", f"piano/ciclo-{cycle['number']}/variante-{variant['number']}/index.html",
                title=f"{cycle['month']} · Variante {variant['number']}", nav="plan", page_id="variant",
                cycle=cycle, variant=variant, prev_variant=prev_variant, next_variant=next_variant,
                breadcrumbs=[
                    {"label": "Panoramica", "path": "index.html"},
                    {"label": "Piano", "path": "piano/index.html"},
                    {"label": f"Ciclo {cycle['number']}", "path": f"piano/ciclo-{cycle['number']}/index.html"},
                    {"label": f"Variante {variant['number']}"},
                ],
            )
            for day in variant["days"]:
                relative_path = f"piano/ciclo-{cycle['number']}/variante-{variant['number']}/{day['d_code'].lower()}/index.html"
                output_dir = (DOCS / relative_path).parent

                def adjacent(target_number: int) -> dict[str, str] | None:
                    target = day_lookup.get(target_number)
                    if not target:
                        return None
                    target_path = DOCS / f"piano/ciclo-{target['cycle']}/variante-{target['variant']}/{target['d_code'].lower()}/index.html"
                    return {
                        "href": os.path.relpath(target_path, output_dir).replace(os.sep, "/"),
                        "label": f"{day_short(target['d_code'])} · giorno {target['global_day']}",
                    }

                render(
                    "day.html", relative_path, title=f"Giorno {day['global_day']} · {day_short(day['d_code'])}", nav="plan", page_id="day",
                    cycle=cycle, variant=variant, day=day, prev_day=adjacent(day["global_day"] - 1), next_day=adjacent(day["global_day"] + 1),
                    breadcrumbs=[
                        {"label": "Panoramica", "path": "index.html"},
                        {"label": "Piano", "path": "piano/index.html"},
                        {"label": f"Ciclo {cycle['number']}", "path": f"piano/ciclo-{cycle['number']}/index.html"},
                        {"label": f"Variante {variant['number']}", "path": f"piano/ciclo-{cycle['number']}/variante-{variant['number']}/index.html"},
                        {"label": day_short(day["d_code"])},
                    ],
                )

    recipes_sorted = sorted(recipes.values(), key=lambda item: item["title"].casefold())
    meal_types = sorted({meal_type for recipe in recipes_sorted for meal_type in recipe["meal_types"]}, key=str.casefold)
    cuisines = sorted({cuisine for recipe in recipes_sorted for cuisine in recipe["cuisines"]}, key=str.casefold)
    render(
        "recipes_index.html", "ricette/index.html", title="Ricette", nav="recipes", page_id="recipes",
        recipes=recipes_sorted, meal_types=meal_types, cuisines=cuisines,
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Ricette"}],
    )
    for recipe in recipes_sorted:
        render(
            "recipe.html", f"ricette/{recipe['slug']}/index.html", title=recipe["title"], nav="recipes", page_id="recipe", recipe=recipe,
            breadcrumbs=[
                {"label": "Panoramica", "path": "index.html"},
                {"label": "Ricette", "path": "ricette/index.html"},
                {"label": recipe["title"]},
            ],
        )

    render(
        "shopping_range.html", "spesa/index.html", title="Spesa per date", nav="shopping", page_id="shopping-range", cycles_href="cicli/index.html",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Spesa"}],
    )
    render(
        "shopping_range.html", "spesa/intervallo/index.html", title="Spesa per date", nav="shopping", page_id="shopping-range", cycles_href="../cicli/index.html",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Spesa", "path": "spesa/index.html"}, {"label": "Intervallo di date"}],
    )
    render(
        "shopping_index.html", "spesa/cicli/index.html", title="Spesa per ciclo e variante", nav="shopping", page_id="shopping", cycles=cycles, list_prefix="../",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Spesa", "path": "spesa/index.html"}, {"label": "Cicli e varianti"}],
    )
    for cycle in cycles:
        cycle_items = shopping_cycles.get(cycle["number"], [])
        render(
            "shopping_list.html", f"spesa/ciclo-{cycle['number']}/index.html", title=f"Spesa · {cycle['month']}", nav="shopping", page_id="shopping-list",
            heading=f"Ciclo {cycle['number']} · {cycle['month']}", subheading="Riepilogo indicativo dei 30 giorni; gli alimenti freschi vanno acquistati per variante.",
            items=cycle_items, grouped_items=group_by_category(cycle_items), list_key=f"ciclo-{cycle['number']}",
            plan_href=f"../../piano/ciclo-{cycle['number']}/index.html",
            breadcrumbs=[
                {"label": "Panoramica", "path": "index.html"},
                {"label": "Spesa", "path": "spesa/index.html"},
                {"label": f"Ciclo {cycle['number']} · {cycle['month']}"},
            ],
        )
        for variant in cycle["variants"]:
            items = shopping_variants.get((cycle["number"], variant["number"]), [])
            render(
                "shopping_list.html", f"spesa/ciclo-{cycle['number']}/variante-{variant['number']}/index.html",
                title=f"Spesa · {cycle['month']} · Variante {variant['number']}", nav="shopping", page_id="shopping-list",
                heading=f"{cycle['month']} · Variante {variant['number']}",
                subheading=f"Acquisti per i giorni {variant['start_day']}–{variant['end_day']} della matrice G–N–SN–R1–R2.",
                items=items, grouped_items=group_by_category(items), list_key=f"ciclo-{cycle['number']}-variante-{variant['number']}",
                plan_href=f"../../../piano/ciclo-{cycle['number']}/variante-{variant['number']}/index.html",
                breadcrumbs=[
                    {"label": "Panoramica", "path": "index.html"},
                    {"label": "Spesa", "path": "spesa/index.html"},
                    {"label": f"Ciclo {cycle['number']}", "path": f"spesa/ciclo-{cycle['number']}/index.html"},
                    {"label": f"Variante {variant['number']}"},
                ],
            )

    render(
        "download.html", "download/index.html", title="Download", nav="downloads", page_id="downloads",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Download"}],
    )
    completed = [
        "Importazione verificata del file Excel revisionato",
        "Pagine statiche per 6 cicli, 36 varianti e 180 giorni",
        "Calendario civile e pagina Oggi con gestione della coda N",
        "Finestra mobile completa di 48 ore, distinta in 0-24 e 24-48 ore",
        "Lista della spesa per intervallo con aggregazione di 3.190 record ingrediente-giorno",
        "Ricerca globale, filtri ricette ed esportazione ICS",
        "Backup locale, stampa e checklist salvate nel browser",
        "PWA installabile con manifest, service worker e indicatori online/offline",
        "Pacchetto offline opzionale per pagine, ricette, piano e dati",
        "Strategia di aggiornamento cache con notifica e attivazione esplicita",
        "Redesign UI completo, responsive e accessibile, con asset grafici locali leggeri",
        "Test automatici di date, spesa, 48 ore, ricerca, ICS, PWA e collegamenti",
        "Studio ingredienti V5 con catalogo base/personale, revisioni, conversioni, validazioni e archiviazione",
        "Studio ricette V5 con ingredient picker, ricalcolo nutrizionale, porzioni, duplicazione e versionamento",
        "Calendario effettivo V5 modificabile, aderenza, turni personalizzati e undo/redo",
        "Compositore della giornata V5 con ricette base/personali, porzioni, blocchi e suggerimenti locali",
        "Resolver unico del piano effettivo V5 per Home, Oggi, preparazioni, spesa, ricerca ed export ICS",
        "Release 5.1.0: nuova nomenclatura e colori dei turni, Mattino/Pomeriggio, Gestisci giornata e preferenze alimentari locali",
        "Release 5.2.1: correzione recupero calendario personale e vista Oggi più compatta",
    ]
    todo = [
        "Verifica manuale dell’installazione e del comportamento standalone su dispositivi Android e iOS reali",
        "Backlog V5: preferiti, note locali e sostituzioni equivalenti interattive",
        "Backlog V5: pianificazione batch-cooking e dispensa personale",
    ]
    decisions = [
        {"title": "Hosting", "text": "Frontend statico pubblicabile dalla cartella docs su GitHub Pages, senza backend; IndexedDB conserva localmente il piano personale."},
        {"title": "Fonte dati", "text": "L'Excel revisionato resta la fonte autorevole del piano base; HTML e JSON vengono rigenerati e IndexedDB applica il livello personale effettivo."},
        {"title": "Asset", "text": "CSS, JavaScript, icone e illustrazioni vivono in static e vengono copiati nella build; non si modificano i file generati."},
        {"title": "Privacy", "text": "Nessun dato personale viene pubblicato. Piano, ricette, ingredienti, modifiche e preferenze restano nel browser e sono esportabili in JSON."},
        {"title": "UI", "text": "Palette berry, blush, teal e lilla; grafica leggera con illustrazioni SVG locali e font di sistema per evitare rallentamenti."},
        {"title": "Accessibilità", "text": "Colori sempre accompagnati da sigle e testo, focus visibile, contrasto controllato e supporto a movimento ridotto."},
        {"title": "Calendario", "text": "Il template parte da C1/V1/G; il calendario effettivo può poi contenere eccezioni, inserimenti, rimozioni, Mattino, Pomeriggio, turni personalizzati e pasti personalizzati."},
        {"title": "48 ore", "text": "La finestra è mobile e inclusiva: dal riferimento fino a +48 ore, visualizzata in blocchi 0-24 e 24-48."},
        {"title": "Offline", "text": "Le risorse essenziali sono memorizzate automaticamente; l'intera libreria statica è scaricabile facoltativamente dalle Utilità."},
        {"title": "Aggiornamenti", "text": "Una nuova build usa cache versionate; l'utente riceve un avviso e decide quando attivare la versione pronta."},
        {"title": "Stato locale", "text": "Configurazione semplice in localStorage e dati V5 strutturati in IndexedDB; ingredienti e ricette personali restano sul dispositivo e sono esportabili in JSON."},
        {"title": "Ricette V5", "text": "Le ricette base sono immutabili; le personali hanno versioni successive e ogni riga conserva la revisione esatta dell'ingrediente usata nel calcolo."},
        {"title": "Piano effettivo", "text": "Home, Oggi, 48 ore, spesa, ricerca e ICS risolvono lo stesso piano locale: date civili, versioni ricetta, porzioni e pasti effettivamente assegnati."},
        {"title": "ICS", "text": "L'export usa il piano effettivo nel fuso Europe/Rome; G, N e i turni personalizzati con orario sono eventi orari, gli altri tipi di giornata sono eventi giornalieri."},
    ]
    render(
        "project.html", "progetto/index.html", title="Stato del progetto", nav="", page_id="project",
        completed=completed, todo=todo, decisions=decisions,
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Progetto"}],
    )
    render("404.html", "404.html", title="Pagina non trovata", nav="", page_id="404")
    render(
        "offline.html", "offline/index.html", title="Modalità offline", nav="tools", page_id="offline",
        breadcrumbs=[{"label": "Panoramica", "path": "index.html"}, {"label": "Modalità offline"}],
    )

    public_cycles = []
    for cycle in cycles:
        public_cycle = {key: cycle[key] for key in ("number", "month", "season_note", "avg", "flexible_count")}
        public_cycle["variants"] = cycle["variants"]
        public_cycles.append(public_cycle)
    build_time = datetime.now(timezone.utc).isoformat()
    write_json(DOCS / "data" / "plan.json", {"version": VERSION, "generated_at": build_time, "summary": overall_summary, "cycles": public_cycles})
    write_json(DOCS / "data" / "recipes.json", {"version": VERSION, "generated_at": build_time, "recipes": recipes_sorted})
    write_json(
        DOCS / "data" / "shopping.json",
        {
            "version": VERSION, "generated_at": build_time,
            "variants": {f"c{cycle}-v{variant}": items for (cycle, variant), items in shopping_variants.items()},
            "cycles": {f"c{cycle}": items for cycle, items in shopping_cycles.items()},
        },
    )
    calendar_days = []
    for day in days:
        calendar_days.append(
            {
                "global_day": day["global_day"],
                "cycle": day["cycle"],
                "month": day["month"],
                "variant": day["variant"],
                "day_in_variant": day["day_in_variant"],
                "d_code": day["d_code"],
                "shift_name": day["shift_name"],
                "shift_hours": day["shift_hours"],
                "flexible": day["flexible"],
                "total": day["total"],
                "path": f"piano/ciclo-{day['cycle']}/variante-{day['variant']}/{day['d_code'].lower()}/index.html",
                "meals": [
                    {
                        "time": meal["time"],
                        "minute_of_day": meal["minute_of_day"],
                        "day_offset": meal["day_offset"],
                        "meal_type": meal["meal_type"],
                        "title": meal["title"],
                        "kcal": meal["kcal"],
                        "prep_minutes": meal["prep_minutes"],
                        "prepare_ahead": meal["prepare_ahead"],
                        "cold": meal["cold"],
                        "reheat": meal["reheat"],
                        "fridge": meal["fridge"],
                        "ingredients": meal["ingredients"],
                        "cuisine": meal["cuisine"],
                        "notes": meal["notes"],
                        "flexible": meal["flexible"],
                        "recipe_slug": meal["recipe_slug"],
                        "anchor": meal["anchor"],
                    }
                    for meal in day["meals"]
                ],
            }
        )
    write_json(
        DOCS / "data" / "calendar.json",
        {
            "version": VERSION,
            "generated_at": build_time,
            "duration_days": 180,
            "start_mapping": {"cycle": 1, "variant": 1, "d_code": "D1", "global_day": 1},
            "days": calendar_days,
        },
    )

    ingredient_day_rows = [
        {"global_day": day_number, "ingredients": ingredient_days.get(day_number, [])}
        for day_number in range(1, 181)
    ]
    write_json(
        DOCS / "data" / "shopping-range.json",
        {
            "version": VERSION,
            "generated_at": build_time,
            "categories": CATEGORY_ORDER,
            "days": ingredient_day_rows,
            "rules": rounding_rules,
        },
    )
    search_entries = build_search_index(cycles, days, recipes_sorted, ingredient_catalog)
    write_json(
        DOCS / "data" / "search-index.json",
        {"version": VERSION, "generated_at": build_time, "entries": search_entries},
    )

    # V5 Phase 2: publish immutable seed data needed for local IndexedDB initialization.
    v5_out = DOCS / "data" / "v5"
    v5_out.mkdir(parents=True, exist_ok=True)
    for source_name, target_name in [
        ("base-dataset-manifest.json", "base-dataset-manifest.json"),
        ("ingredients.base.v1.json", "ingredients.base.v1.json"),
        ("recipes.base.v1.json", "recipes.base.v1.json"),
        ("plan-template.base.v1.json", "plan-template.base.v1.json"),
    ]:
        shutil.copy2(ROOT / "v5_data" / "base" / source_name, v5_out / target_name)

    write_json(
        DOCS / "data" / "build-meta.json",
        {
            "version": VERSION,
            "generated_at": build_time,
            "source_file": SOURCE_XLSX.name,
            "source_sha256": sha256(SOURCE_XLSX),
            "counts": counts,
        },
    )
    offline_manifest = build_offline_manifest()

    print(json.dumps({
        "status": "ok",
        "version": VERSION,
        "counts": counts,
        "recipes": len(recipes_sorted),
        "offline_assets": offline_manifest["asset_count"],
        "offline_bytes": offline_manifest["total_bytes"],
        "output": str(DOCS),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
